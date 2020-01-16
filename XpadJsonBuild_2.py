# coding=utf-8
import datetime
import json
import os
from tkinter import filedialog
from tkinter import ttk

import openpyxl

from BaseXpadJsonBuild import *


class XpadJsonBuild2(BaseXpadJsonBuild):
    excelPath = ""

    mAdResultMap = {}

    filename = None
    path = None
    listBox: Listbox = None

    def selectPath(self):
        self.filename = filedialog.askopenfilename(filetypes=[("excel格式", "xlsx")])
        self.path.set(self.filename)
        self.insertListBoxMessage("选择路径 :" + self.filename)

    def startCreateJson(self):
        self.listBox.delete(0, END)
        self.startAndBuild(self.filename)
        pass

    def addChannelIds(self, slot: list, adSheet):
        maxColum = adSheet.max_row
        channels = {}
        channelDatas = []
        adPriorityList = []

        for sidRowIndex in range(1, maxColum):

            currentIndex = sidRowIndex + 1
            sidValue = self.getCloumeValueColumValue(currentIndex, "sid", adSheet)

            if sidValue is not None:
                channels = {}
                channelDatas = []
                adPriorityList = []
                channels["channels"] = channelDatas
                slot.append(channels)
                channels["sid"] = sidValue
            else:
                # 跳过不属于当前单元格的输出
                merged_value = self.get_merged_cells_value(adSheet, currentIndex, 4)
                if merged_value != channels["sid"]:
                    continue

            platformValue = self.getCloumeValueColumValue(currentIndex, "Platform", adSheet)
            if platformValue is None:
                self.insertListBoxMessage(
                    "Platform is None --- sid = " + str(channels["sid"]) + "----- 行 = " + str(currentIndex))
                logging.warning("Platform is None --- sid = " + str(channels["sid"]) + "----- 行 = " + str(currentIndex))
                continue

            adSourceIdValue = self.getCloumeValueColumValue(currentIndex, "广告ID", adSheet)
            if adSourceIdValue is None:
                self.insertListBoxMessage(
                    "广告ID is None --- sid = " + str(channels["sid"]) + "----- 行 = " + str(currentIndex))
                logging.warning("广告ID is None --- sid = " + str(channels["sid"]) + "----- 行 = " + str(currentIndex))
                continue

            adWatingTimeValue = self.getCloumeValueColumValue(currentIndex, "广告超时时间", adSheet)
            if adWatingTimeValue is None:
                adWatingTimeValue = 4000
            elif adWatingTimeValue < 1000:
                self.insertListBoxMessage(
                    "广告超时时间 是毫秒的 =" + str(adSourceIdValue) + "--对应 = " + adWatingTimeValue + " 是错误的")
                logging.error("广告超时时间 是毫秒的 =" + str(adSourceIdValue) + "--对应 = " + adWatingTimeValue + " 是错误的")

            adExpriedTImeValue = self.getCloumeValueColumValue(currentIndex, "广告过期时间", adSheet)
            if adExpriedTImeValue is None:
                adExpriedTImeValue = 40
            elif adExpriedTImeValue < 0:
                self.insertListBoxMessage(
                    "广告过期时间 是分钟的 =" + str(adSourceIdValue) + "--对应 = " + adWatingTimeValue + " 是错误的")
                logging.error("广告过期时间 是分钟的 =" + str(adSourceIdValue) + "--对应 = " + adWatingTimeValue + " 是错误的")

            adDetailsValue = self.getCloumeValueColumValue(currentIndex, "备注", adSheet)
            if adDetailsValue is None:
                self.insertListBoxMessage("备注内容为None  跳过= " + str(adSourceIdValue))
                logging.warning("备注内容为None  跳过= " + str(adSourceIdValue))
                continue

            adExtraType = self.getAdExtraTypeValue(platformValue, adDetailsValue, adSourceIdValue)

            self.checkPlatformValueData(platformValue, adSourceIdValue)

            channelItem = {}
            channelItemExtra = {}

            adPriorityValue = self.getCloumeValueColumValue(currentIndex, "广告优先级", adSheet)

            if adPriorityValue is not None:
                adPriorityListlength = len(adPriorityList)

                if adPriorityListlength == 0:
                    channelDatas.append(channelItem)
                    adPriorityList.append(adPriorityValue)
                else:
                    for i in range(0, adPriorityListlength):
                        if adPriorityValue <= adPriorityList[i]:
                            adPriorityList.insert(i, adPriorityValue)
                            channelDatas.insert(i, channelItem)
                            break
                        else:
                            # i+1 小于最长的长度，可以再向前进一格子
                            if (i + 1) < adPriorityListlength:
                                continue
                            else:
                                # i+1 超出了长度 数据等于排序长度
                                if len(channelDatas) <= adPriorityListlength:
                                    adPriorityList.append(adPriorityValue)
                                    channelDatas.append(channelItem)
                                else:
                                    adPriorityList.append(adPriorityValue)
                                    channelDatas.insert(i + 1, channelItem)

            else:
                channelDatas.append(channelItem)

            logging.debug("优先级顺序 -- " + str(adPriorityList))

            channelItem["extra"] = channelItemExtra
            channelItem["channel"] = platformValue
            channelItem["pid"] = adSourceIdValue
            channelItem["wt"] = adWatingTimeValue
            channelItem["ttl"] = adExpriedTImeValue

            channelItemExtra["type"] = adExtraType
            if adExtraType == AD_TYPE_CSJ_PERSONAL_PLATE_INTERSTITIAL:
                channelItemExtra["image_ratio"] = "2:3"
            elif adExtraType == AD_TYPE_CSJ_PERSONAL_PLATE_FEED:
                channelItemExtra["dip_w"] = "300"
                channelItemExtra["dip_h"] = "250"
            elif adExtraType == AD_TYPE_YLH_PERSONAL_PLATE_FEED:
                channelItemExtra["pixel_w"] = "300"
                channelItemExtra["pixel_h"] = "250"

        self.findTitleInColum("广告位名称", adSheet)

        pass

    def main(self, excelPath):
        # 开始读取
        wb = openpyxl.load_workbook(excelPath)
        sheetNames = wb.sheetnames
        adSheet: Worksheet = wb[str(sheetNames[0])]

        maxColumn = int(adSheet.max_column)
        self.insertListBoxMessage("最大列 = " + str(maxColumn))
        self.insertListBoxMessage("最大行 = " + str(adSheet.max_row))
        logging.info("最大列 = " + str(maxColumn))
        logging.info("最大行 = " + str(adSheet.max_row))

        appLicenseId = self.getTitleColumValue("app license id", adSheet)
        csjAppId = self.getTitleColumValue("穿山甲应用ID", adSheet)
        ylhAppId = self.getTitleColumValue("广点通应用ID", adSheet)
        kshAppId = self.getTitleColumValue("快手应用ID", adSheet)
        appId = self.getTitleColumValue("appId", adSheet)

        self.mAdResultMap["ls"] = appLicenseId
        self.mAdResultMap["rt"] = 20
        adResultDataMap = {}
        self.mAdResultMap["data"] = adResultDataMap

        adResultDataMap["ls"] = appLicenseId
        adResultDataMap["csj"] = csjAppId
        adResultDataMap["ylh"] = ylhAppId
        adResultDataMap["ksh"] = kshAppId
        adResultDataMap["appid"] = appId
        slot = []
        adResultDataMap["slot"] = slot

        self.addChannelIds(slot, adSheet)

        self.mAdResultMap["status"] = 1

    def startAndBuild(self, excelPath):
        logging.basicConfig(level=logging.INFO)
        self.insertListBoxMessage("XPAD 2.0 json脚本生成工具")
        logging.info("XPAD 2.0 json脚本生成工具")

        if not os.path.exists(excelPath):
            self.insertListBoxMessage("需要解析的Excel文件 不存在")
            logging.error("需要解析的Excel文件 不存在")
            self.insertListBoxMessage("解析将会停止")

        fileSub = os.path.splitext(os.path.basename(excelPath))[1]
        if fileSub != ".xlsx":
            self.insertListBoxMessage("请输入正确的Excel文件->" + fileSub)
            logging.error("请输入正确的Excel文件->" + fileSub)
            self.insertListBoxMessage("解析将会停止")
        self.insertListBoxMessage("请确保需要解析的广告数据表在第一个...")
        self.insertListBoxMessage("开始生成 ...")
        logging.warning("请确保需要解析的广告数据表在第一个...")
        self.insertListBoxMessage("开始生成 ..." + str(datetime.datetime.now()))

        self.main(excelPath)

        jsonResult = json.dumps(self.mAdResultMap, indent=4, ensure_ascii=False)
        fileName = os.path.splitext(os.path.basename(excelPath))[0]
        fileDir = os.path.dirname(excelPath)
        resultAdFileJson = os.path.join(fileDir, fileName + "_xpad_ad_release_2.0.json")
        self.insertListBoxMessage(resultAdFileJson)
        logging.info(resultAdFileJson)

        if os.path.exists(resultAdFileJson):
            resultAdFile = open(resultAdFileJson, 'w')
            resultAdFile.write(jsonResult)
            resultAdFile.close()
        else:
            resultAdFile = open(resultAdFileJson, 'a')
            resultAdFile.write(jsonResult)
            resultAdFile.close()

        self.insertListBoxMessage("生成成功")
        logging.info("生成成功")

    def creatMainUi(self):

        root = Tk()
        root.title("XPAD 2.0 json脚本生成工具")
        root.geometry("1000x618")
        root.resizable(False, False)

        self.path = StringVar()

        topFrame = Frame(root)
        topFrame.pack(side=TOP)

        Label(topFrame, text="目标路径:").pack(side=LEFT, padx=5, pady=10)
        ttk.Entry(topFrame, textvariable=self.path).pack(side=LEFT, padx=5, pady=10)
        ttk.Button(topFrame, text="路径选择", command=self.selectPath).pack(side=LEFT, padx=5, pady=10)
        ttk.Button(topFrame, text="开始生成", command=self.startCreateJson).pack(side=LEFT, padx=5, pady=10)

        bottomFrame = Frame(root)
        scrollbar = Scrollbar(bottomFrame)
        scrollbar.pack(side=RIGHT, fill=Y)
        self.listBox = Listbox(root, yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.listBox.yview)
        self.listBox.pack(side=LEFT, fill=BOTH, expand=YES, pady=10)
        bottomFrame.pack(side=TOP, fill=BOTH, expand=YES)

        self.insertListBoxMessage("欢迎来到XPAD 2.0 json脚本生成工具")
        self.insertListBoxMessage("请选择路径 :")

        root.mainloop()


if __name__ == '__main__':
    argLen = len(sys.argv)
    build_ = XpadJsonBuild2()
    if argLen < 2:
        build_.creatMainUi()
    else:
        excelPath = sys.argv[1]
        build_.startAndBuild(excelPath)
