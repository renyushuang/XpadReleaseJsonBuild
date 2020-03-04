# coding=utf-8
import datetime
import json
import os
from tkinter import filedialog
from tkinter import ttk

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from BaseXpadJsonBuild import *

gdtReward = "gdtReward"
# 广点通激励视频
gdtFeed = "gdtFeed"
# 广点通信息流
gdtInter = "gdtInter"
# 广点通插屏( dialog))
gdtSplash = "gdtSplash"
# 广点通开屏/闪屏
gdtExpress = "gdtExpress"
# 广点通原生模板

csjReward = "csjReward"
# 穿山甲激励视频
csjSplash = "csjSplash"
# 穿山甲开屏/闪屏
csjFeed = "csjFeed"
# 穿山甲信息流
csjVideo = "csjVideo"
# 穿山甲全屏视频
csjNative = "csjNative"
# 穿山甲原生/原生插屏
csjInter = "csjInter"
# 穿山甲插屏( dialog))
csjNativeExpress = "csjNativeExpress"
# 穿山甲个性化模板信息流


ksReward = "ksReward"
# 快手激励视频
ksNative = "ksNative"
# 手原生/信息
ksVideo = "ksVideo"
# 快手全屏视频


baiduFeed = "baiduFeed"
# 百度信息流(baiduFeed)，
baiduReward = "baiduReward"
# 百度激励视频(baiduReward)，
baiduSplash = "baiduSplash"
# 百度闪屏(baiduSplash)


AD_TYPE_BAIDU_NAME = "百度"
AD_TYPE_BAIDU = "baidu"

AD_TYPE_YLH_NAME = "广点通"
AD_TYPE_YLH = "ylh"

AD_TYPE_CSJ_NAME = "穿山甲"
AD_TYPE_CSJ = "csj"

AD_TYPE_KSH_NAME = "快手"
AD_TYPE_KSH = "ksh"

# 广告备注命名格式 广告源名称-以下关键字 例如：穿山甲-开屏、百度-信息流、穿山甲-原生模版 穿山甲-原生-自渲染等
# 开屏
AD_TYPE_OPEN = "开屏"
AD_TYPE_SPLASH = "闪屏"

# 模版原生
AD_TYPE_NATIVE_PLATE = "原生模版"
AD_TYPE_PLATE_RENDER = "模版渲染"
AD_TYPE_PLATE = "模版"
AD_TYPE_PLATE_INFO = "个性化模板信息流"
# 原生
AD_TYPE_NATIVE = "原生"
AD_TYPE_NATIVE_FEED = "自渲染"
AD_TYPE_NATIVE_SELF = "自渲染原生"
AD_TYPE_INFO = "信息流"
AD_TYPE_INFO_SELF = "自渲染信息流"
# 插屏
AD_TYPE_INTERSTITIAL = "插屏"
AD_TYPE_INTERSTITIAL_PLATE = "个性化模版插屏"

# 全屏视频
AD_TYPE_FULL_SCREEN_VIDEO = "全屏视频"

# 全屏视频
AD_TYPE_BANNER = "banner"
# 激励视频
AD_TYPE_REWARD = "激励视频"


class CoinSdkJsonBuildV3(BaseXpadJsonBuild):
    excelPath = ""

    mAdResultMap = {}

    filename = None
    path = None
    listBox: Listbox = None

    def selectPath(self):
        self.filename = filedialog.askopenfilename(filetypes=[("excel格式", "xlsx")])
        self.path.set(self.filename)
        self.insertListBoxMessage("选择路径 :" + self.filename)

    def startCreateJson(self):
        self.listBox.delete(0, END)
        self.startAndBuild(self.filename)
        pass

    def hashSplashTypeString(self, adDetailsValuelist):
        if self.hasAdTypeString(adDetailsValuelist, AD_TYPE_OPEN):
            return True
        if self.hasAdTypeString(adDetailsValuelist, AD_TYPE_SPLASH):
            return True
        return False

    def hasPlateNative(self, adDetailsValuelist):
        if self.hasAdTypeString(adDetailsValuelist, AD_TYPE_NATIVE_PLATE):
            return True
        if self.hasAdTypeString(adDetailsValuelist, AD_TYPE_PLATE):
            return True
        if self.hasAdTypeString(adDetailsValuelist, AD_TYPE_PLATE_RENDER):
            return True
        if self.hasAdTypeString(adDetailsValuelist, AD_TYPE_PLATE_INFO):
            return True
        return False

    def hasNative(self, adDetailsValuelist):
        if self.hasAdTypeString(adDetailsValuelist, AD_TYPE_NATIVE):
            return True
        if self.hasAdTypeString(adDetailsValuelist, AD_TYPE_NATIVE_FEED):
            return True
        if self.hasAdTypeString(adDetailsValuelist, AD_TYPE_NATIVE_SELF):
            return True
        if self.hasAdTypeString(adDetailsValuelist, AD_TYPE_INFO):
            return True
        if self.hasAdTypeString(adDetailsValuelist, AD_TYPE_INFO_SELF):
            return True
        return False

    def getAdExtraTypeValue(self, platformValue, adDetailsValue, adSourceIdValue):
        adDetailsValuelist: list = adDetailsValue.split("-")
        # if len(adDetailsValuelist) < 2:
        #     self.insertListBoxMessage("备注 命名错误 -- " + str(adSourceIdValue))
        #     logging.error("备注 命名错误 -- " + str(adSourceIdValue))

        if platformValue == AD_TYPE_CSJ:
            if self.hashSplashTypeString(adDetailsValuelist):
                return csjSplash
            elif self.hasAdTypeString(adDetailsValuelist, AD_TYPE_INTERSTITIAL) \
                    or self.hasAdTypeString(adDetailsValuelist, AD_TYPE_INTERSTITIAL_PLATE):
                return csjInter
            elif self.hasAdTypeString(adDetailsValuelist, AD_TYPE_FULL_SCREEN_VIDEO):
                return csjVideo
            elif self.hasPlateNative(adDetailsValuelist):
                return csjNativeExpress
            elif self.hasAdTypeString(adDetailsValuelist, AD_TYPE_NATIVE_FEED):
                return csjFeed
            elif self.hasAdTypeString(adDetailsValuelist, AD_TYPE_NATIVE) \
                    or self.hasAdTypeString(adDetailsValuelist, AD_TYPE_INFO_SELF):
                return csjNative
            elif self.hasAdTypeString(adDetailsValuelist, AD_TYPE_REWARD):
                return csjReward
            else:
                self.insertListBoxMessage("不支持这种类型 -- " + adDetailsValue + " -- 穿山甲的广告id为 = " + str(adSourceIdValue))
                logging.error("不支持这种类型 -- " + adDetailsValue + " -- 穿山甲的广告id为 = " + str(adSourceIdValue))

        elif platformValue == AD_TYPE_YLH:
            if self.hashSplashTypeString(adDetailsValuelist):
                return gdtSplash
            elif self.hasNative(adDetailsValuelist):
                return gdtFeed
            elif self.hasAdTypeString(adDetailsValuelist, AD_TYPE_REWARD):
                return gdtReward
            elif self.hasAdTypeString(adDetailsValuelist, AD_TYPE_INTERSTITIAL) \
                    or self.hasAdTypeString(adDetailsValuelist, AD_TYPE_INTERSTITIAL_PLATE):
                return gdtInter
            elif self.hasPlateNative(adDetailsValuelist):
                return gdtExpress
            else:
                self.insertListBoxMessage("不支持这种类型 --  " + adDetailsValue + "-- 广点通id为 = " + str(adSourceIdValue))
                logging.error("不支持这种类型 --  " + adDetailsValue + "-- 广点通id为 = " + str(adSourceIdValue))
            pass
        elif platformValue == AD_TYPE_KSH:
            if self.hasNative(adDetailsValuelist):
                return ksNative
            elif self.hasAdTypeString(adDetailsValuelist, AD_TYPE_FULL_SCREEN_VIDEO) \
                    or self.hasAdTypeString(adDetailsValuelist, AD_TYPE_INTERSTITIAL) \
                    or self.hasAdTypeString(adDetailsValuelist, AD_TYPE_INTERSTITIAL_PLATE):
                return ksVideo
            elif self.hasAdTypeString(adDetailsValuelist, AD_TYPE_REWARD):
                return ksReward
            else:
                self.insertListBoxMessage("不支持这种类型 -- " + adDetailsValue + "-- 快手id为 = " + str(adSourceIdValue))
                logging.error("不支持这种类型 -- " + adDetailsValue + "-- 快手id为 = " + str(adSourceIdValue))
            pass
        elif platformValue == AD_TYPE_BAIDU:
            if self.hashSplashTypeString(adDetailsValuelist):
                return baiduSplash
            elif self.hasNative(adDetailsValuelist):
                return baiduFeed
            elif self.hasAdTypeString(adDetailsValuelist, AD_TYPE_REWARD):
                return baiduReward
            else:
                self.insertListBoxMessage("不支持这种类型 -- " + adDetailsValue + "-- 快手id为 = " + str(adSourceIdValue))
                logging.error("不支持这种类型 -- " + adDetailsValue + "-- 快手id为 = " + str(adSourceIdValue))

            pass
        else:
            self.insertListBoxMessage("没有这个类型 广告id是 = " + str(adSourceIdValue))
            logging.error("没有这个类型 广告id是 = " + str(adSourceIdValue))
            self.insertListBoxMessage("解析将会停止")

        self.insertListBoxMessage("不支持这种类型 -- " + adDetailsValue + "-- id为 = " + str(adSourceIdValue))
        logging.error("不支持这种类型 -- " + adDetailsValue + "-- id为 = " + str(adSourceIdValue))
        return None

    def addChannelIds(self, slot: dict, adSheet):
        maxColum = adSheet.max_row

        channelDatas = []
        adPriorityList = []

        for sidRowIndex in range(1, maxColum):

            currentIndex = sidRowIndex + 1
            sidValue = self.getCloumeValueColumValue(currentIndex, "sid", adSheet)

            if sidValue is not None:
                channelDatas = []
                adPriorityList = []
                slot[sidValue] = channelDatas
            else:
                pass
                # 跳过不属于当前单元格的输出

                merged_value = self.get_merged_cells_value(adSheet, currentIndex, self.findTitleInColum("sid", adSheet))
                print("merged_value" + str(merged_value))
                if not slot.__contains__(merged_value):
                    continue

            merged_value = self.get_merged_cells_value(adSheet, currentIndex, self.findTitleInColum("sid", adSheet))

            platformValue = self.getCloumeValueColumValue(currentIndex, "广告渠道", adSheet)
            if platformValue is None:
                self.insertListBoxMessage(
                    "Platform is None --- sid = " + str(merged_value) + "----- 行 = " + str(currentIndex))
                logging.warning("Platform is None --- sid = " + str(merged_value) + "----- 行 = " + str(currentIndex))
                continue
            if platformValue == "穿山甲":
                platformValue = "csj"
            elif platformValue == "广点通":
                platformValue = "ylh"
            elif platformValue == "快手":
                platformValue = "ksh"
            elif platformValue == "百度":
                platformValue = "baidu"

            adSourceIdValue = self.getCloumeValueColumValue(currentIndex, "广告ID", adSheet)
            if adSourceIdValue is None:
                self.insertListBoxMessage(
                    "广告ID is None --- sid = " + str(merged_value) + "----- 行 = " + str(currentIndex))
                logging.warning("广告ID is None --- sid = " + str(merged_value) + "----- 行 = " + str(currentIndex))
                continue

            adDetailsValue = self.getCloumeValueColumValue(currentIndex, "广告类型", adSheet)
            if adDetailsValue is None:
                self.insertListBoxMessage("广告类型 内容为None  跳过= " + str(adSourceIdValue))
                logging.warning("广告类型 内容为None  跳过= " + str(adSourceIdValue))
                continue

            adExtraType = self.getAdExtraTypeValue(platformValue, adDetailsValue, adSourceIdValue)
            self.checkPlatformValueData(platformValue, adSourceIdValue)

            channelItem = {}

            adPriorityValue = self.getCloumeValueColumValue(currentIndex, "优先级", adSheet)

            if adPriorityValue is not None:
                adPriorityListlength = len(adPriorityList)

                if adPriorityListlength == 0:
                    channelDatas.append(channelItem)
                    adPriorityList.append(adPriorityValue)
                else:
                    for i in range(0, adPriorityListlength):
                        if adPriorityValue <= adPriorityList[i]:
                            adPriorityList.insert(i, adPriorityValue)
                            channelDatas.insert(i, channelItem)
                            break
                        else:
                            # i+1 小于最长的长度，可以再向前进一格子
                            if (i + 1) < adPriorityListlength:
                                continue
                            else:
                                # i+1 超出了长度 数据等于排序长度
                                if len(channelDatas) <= adPriorityListlength:
                                    adPriorityList.append(adPriorityValue)
                                    channelDatas.append(channelItem)
                                else:
                                    adPriorityList.append(adPriorityValue)
                                    channelDatas.insert(i + 1, channelItem)

            else:
                channelDatas.append(channelItem)

            logging.debug("优先级顺序 -- " + str(adPriorityList))

            channelItem["id"] = adSourceIdValue
            channelItem["wt"] = 3000
            channelItem["type"] = adExtraType

        pass

    def main(self, excelPath):
        # 开始读取
        wb = openpyxl.load_workbook(excelPath)
        sheetNames = wb.sheetnames
        adSheet: Worksheet = wb[str(sheetNames[0])]

        maxColumn = int(adSheet.max_column)
        self.insertListBoxMessage("最大列 = " + str(maxColumn))
        self.insertListBoxMessage("最大行 = " + str(adSheet.max_row))
        logging.info("最大列 = " + str(maxColumn))
        logging.info("最大行 = " + str(adSheet.max_row))

        csjAppId = self.getTitleColumValue("穿山甲appid", adSheet)
        ylhAppId = self.getTitleColumValue("广点通appid", adSheet)
        kshAppId = self.getTitleColumValue("快手appid", adSheet)
        baiduAppId = self.getTitleColumValue("百度appid", adSheet)

        platform_ids = []
        self.mAdResultMap["platform_ids"] = platform_ids
        platform_ids.append({"id": csjAppId, "type": "csj"})
        platform_ids.append({"id": kshAppId, "type": "ks"})
        platform_ids.append({"id": ylhAppId, "type": "gdt"})
        platform_ids.append({"id": ylhAppId, "type": "baidu"})

        ad_ids_config_v3 = {}
        self.mAdResultMap["ad_ids_config_v3"] = ad_ids_config_v3
        self.addChannelIds(ad_ids_config_v3, adSheet)

    def startAndBuild(self, excelPath):
        logging.basicConfig(level=logging.INFO)
        self.insertListBoxMessage("金币SDK 3.0 json脚本生成工具")
        logging.info("金币SDK 3.0 json脚本生成工具")

        if not os.path.exists(excelPath):
            self.insertListBoxMessage("需要解析的Excel文件 不存在")
            logging.error("需要解析的Excel文件 不存在")
            self.insertListBoxMessage("解析将会停止")

        fileSub = os.path.splitext(os.path.basename(excelPath))[1]
        if fileSub != ".xlsx":
            self.insertListBoxMessage("请输入正确的Excel文件->" + fileSub)
            logging.error("请输入正确的Excel文件->" + fileSub)
            self.insertListBoxMessage("解析将会停止")
        self.insertListBoxMessage("请确保需要解析的广告数据表在第一个...")
        self.insertListBoxMessage("开始生成 ...")
        logging.warning("请确保需要解析的广告数据表在第一个...")
        self.insertListBoxMessage("开始生成 ..." + str(datetime.datetime.now()))

        self.main(excelPath)

        self.mAdResultMap["banners"] = [{
            "action": "VIDEO",
            "des": "最高可获得10000金币!",
            "iconurl": "http://obs.cn-north-1.myhuaweicloud.com/coincommon/sandbox/icon/2019/08/4dda00d25035e3102a599cbb9bc01ae1.png",
            "title": "今日最赚"
        }]
        self.mAdResultMap["video_interval"] = 30
        self.mAdResultMap["lock_config_new"] = {
            "strategy": True,
            "strategy_show": True
        }

        jsonResult = json.dumps(self.mAdResultMap, indent=4, ensure_ascii=False)
        fileName = os.path.splitext(os.path.basename(excelPath))[0]
        fileDir = os.path.dirname(excelPath)
        resultAdFileJson = os.path.join(fileDir, fileName + "_coinsdk_ad_release_2.0.json")
        self.insertListBoxMessage(resultAdFileJson)
        logging.info(resultAdFileJson)

        if os.path.exists(resultAdFileJson):
            resultAdFile = open(resultAdFileJson, 'w')
            resultAdFile.write(jsonResult)
            resultAdFile.close()
        else:
            resultAdFile = open(resultAdFileJson, 'a')
            resultAdFile.write(jsonResult)
            resultAdFile.close()

        self.insertListBoxMessage("生成成功")
        logging.info("生成成功")

    def creatMainUi(self):

        root = Tk()
        root.title("金币SDK 3.0 json脚本生成工具")
        root.geometry("1000x618")
        root.resizable(False, False)

        self.path = StringVar()

        topFrame = Frame(root)
        topFrame.pack(side=TOP)

        Label(topFrame, text="目标路径:").pack(side=LEFT, padx=5, pady=10)
        ttk.Entry(topFrame, textvariable=self.path).pack(side=LEFT, padx=5, pady=10)
        ttk.Button(topFrame, text="路径选择", command=self.selectPath).pack(side=LEFT, padx=5, pady=10)
        ttk.Button(topFrame, text="开始生成", command=self.startCreateJson).pack(side=LEFT, padx=5, pady=10)

        bottomFrame = Frame(root)
        scrollbar = Scrollbar(bottomFrame)
        scrollbar.pack(side=RIGHT, fill=Y)
        self.listBox = Listbox(root, yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.listBox.yview)
        self.listBox.pack(side=LEFT, fill=BOTH, expand=YES, pady=10)
        bottomFrame.pack(side=TOP, fill=BOTH, expand=YES)

        self.insertListBoxMessage("欢迎来到金币SDK 3.0 json脚本生成工具")
        self.insertListBoxMessage("请选择路径 :")

        root.mainloop()

if __name__ == '__main__':
    argLen = len(sys.argv)
    build_ = CoinSdkJsonBuildV3()
    if argLen < 2:
        build_.creatMainUi()
    else:
        excelPath = sys.argv[1]
        build_.startAndBuild(excelPath)
