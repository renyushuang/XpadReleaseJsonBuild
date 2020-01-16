# coding=utf-8
import datetime
import json
import os
from tkinter import filedialog
from tkinter import ttk

import openpyxl

from BaseXpadJsonBuild import *


class XpadJsonBuildDataPip(BaseXpadJsonBuild):
    native_ad_refr_it: IntVar = None
    native_ad_refr_t_u: IntVar = None
    native_ad_new_usr_st = None

    excelPath = ""

    mAdResultMap = {}

    filename = None
    path = None
    listBox: Listbox = None

    def selectPath(self):
        self.filename = filedialog.askopenfilename(filetypes=[("excel格式", "xlsx")])
        self.path.set(self.filename)
        self.insertListBoxMessage("选择路径 :" + self.filename)

    def startCreateJson(self):
        self.listBox.delete(0, END)
        self.startAndBuild(self.filename)
        pass

    def addChannelIds(self, slot: list, adSheet):
        maxColum = adSheet.max_row

        for sidRowIndex in range(1, maxColum):
            currentIndex = sidRowIndex + 1
            sidValue = self.getCloumeValueColumValue(currentIndex, "sid", adSheet)
            native_ad_new_usr_stValue = 253370736000000
            if self.native_ad_new_usr_st is not None:
                native_ad_new_usr_stValue = self.native_ad_new_usr_st.get()

            if sidValue is not None:
                sidDataPipItem = {}
                self.mAdResultMap[str(sidValue)] = sidDataPipItem
                adUnitNameValue: str = self.getCloumeValueColumValue(currentIndex, "广告位名称", adSheet)
                if adUnitNameValue == AD_TYPE_OPEN:
                    sidDataPipItem["ad_sw_o"] = True
                    sidDataPipItem["ad_sw_n"] = True
                    sidDataPipItem["ad_pro_h"] = 0

                    sidDataPipItem["ad_new_usr_st"] = native_ad_new_usr_stValue
                    sidDataPipItem["ad_refr_sw"] = False
                    sidDataPipItem["ad_refr_it"] = 10000
                    sidDataPipItem["ad_refr_t_u"] = 10
                    self.insertListBoxMessage("" + adUnitNameValue + "----" + str(sidDataPipItem["ad_refr_sw"]))
                    logging.info("" + adUnitNameValue + "----" + str(sidDataPipItem["ad_refr_sw"]))
                elif adUnitNameValue.find(AD_TYPE_INTERSTITIAL) > 0:
                    sidDataPipItem["ad_sw_o"] = True
                    sidDataPipItem["ad_sw_n"] = True
                    sidDataPipItem["ad_pro_h"] = 0
                    sidDataPipItem["ad_new_usr_st"] = native_ad_new_usr_stValue
                    sidDataPipItem["ad_refr_sw"] = False
                    sidDataPipItem["ad_refr_it"] = 10000
                    sidDataPipItem["ad_refr_t_u"] = 10
                    self.insertListBoxMessage("" + adUnitNameValue + "----" + str(sidDataPipItem["ad_refr_sw"]))
                    logging.info("" + adUnitNameValue + "----" + str(sidDataPipItem["ad_refr_sw"]))
                elif adUnitNameValue.find(AD_TYPE_NATIVE) > 0:
                    sidDataPipItem["ad_sw_o"] = True
                    sidDataPipItem["ad_sw_n"] = True
                    sidDataPipItem["ad_pro_h"] = 0
                    sidDataPipItem["ad_new_usr_st"] = native_ad_new_usr_stValue
                    sidDataPipItem["ad_refr_sw"] = True

                    ad_refr_it = 10000
                    if self.native_ad_refr_it is not None:
                        ad_refr_it = self.native_ad_refr_it.get()

                    ad_refr_t_u = 10000
                    if self.native_ad_refr_t_u is not None:
                        ad_refr_t_u = self.native_ad_refr_t_u.get()

                    sidDataPipItem["ad_refr_it"] = ad_refr_it
                    sidDataPipItem["ad_refr_t_u"] = ad_refr_t_u
                    self.insertListBoxMessage("" + adUnitNameValue + "----" + str(sidDataPipItem["ad_refr_sw"]))
                    logging.info("" + adUnitNameValue + "----" + str(sidDataPipItem["ad_refr_sw"]))
                else:
                    self.insertListBoxMessage("这个类型不支持" + adUnitNameValue)
                    logging.warning("这个类型不支持" + adUnitNameValue)

            # platformValue = getCloumeValueColumValue(currentIndex, "Platform", adSheet)
            # adSourceIdValue = getCloumeValueColumValue(currentIndex, "广告ID", adSheet)
            # adWatingTimeValue = getCloumeValueColumValue(currentIndex, "广告超时时间", adSheet)
            # adExpriedTImeValue = getCloumeValueColumValue(currentIndex, "广告过期时间", adSheet)
            # adDetailsValue = getCloumeValueColumValue(currentIndex, "备注", adSheet)
            # adExtraType = getAdExtraTypeValue(platformValue, adDetailsValue, adSourceIdValue)
            # adPriorityValue = getCloumeValueColumValue(currentIndex, "广告优先级", adSheet)

        pass

    def main(self, excelPath):
        # 开始读取
        wb = openpyxl.load_workbook(excelPath)
        sheetNames = wb.sheetnames
        adSheet = wb[str(sheetNames[0])]
        maxColumn = int(adSheet.max_column)
        self.insertListBoxMessage("最大列 = " + str(maxColumn))
        self.insertListBoxMessage("最大行 = " + str(adSheet.max_row))
        print("最大列 = " + str(maxColumn))
        print("最大行 = " + str(adSheet.max_row))

        appLicenseId = self.getTitleColumValue("app license id", adSheet)
        csjAppId = self.getTitleColumValue("穿山甲应用ID", adSheet)
        ylhAppId = self.getTitleColumValue("广点通应用ID", adSheet)
        kshAppId = self.getTitleColumValue("快手应用ID", adSheet)
        appId = self.getTitleColumValue("appId", adSheet)

        slot = []

        self.addChannelIds(slot, adSheet)

    def startAndBuild(self, excelPath):
        logging.basicConfig(level=logging.INFO)
        self.insertListBoxMessage("XPAD data pip json脚本生成工具")
        logging.info("XPAD data pip json脚本生成工具")
        argLen = len(sys.argv)
        # if argLen < 2:
        #     logging.error("请输入想要解析的文件")
        #     exit()
        # excelPath = sys.argv[1]

        if not os.path.exists(excelPath):
            self.insertListBoxMessage("需要解析的Excel文件 不存在")

            logging.error("需要解析的Excel文件 不存在")
            exit()

        fileSub = os.path.splitext(os.path.basename(excelPath))[1]
        if fileSub != ".xlsx":
            self.insertListBoxMessage("请输入正确的Excel文件->" + fileSub)
            logging.error("请输入正确的Excel文件->" + fileSub)
            exit()
        self.insertListBoxMessage("请确保需要解析的广告数据表在第一个...")
        logging.warning("请确保需要解析的广告数据表在第一个...")
        self.insertListBoxMessage("开始生成 ..." + str(datetime.datetime.now()))
        logging.info("开始生成 ...")

        self.main(excelPath)

        jsonResult = json.dumps(self.mAdResultMap, indent=4, ensure_ascii=False)
        fileName = os.path.splitext(os.path.basename(excelPath))[0]
        fileDir = os.path.dirname(excelPath)
        resultAdFileJson = os.path.join(fileDir, fileName + "_xpad_ad_release_data_pip.json")
        self.insertListBoxMessage(resultAdFileJson)
        logging.info(resultAdFileJson)

        if os.path.exists(resultAdFileJson):
            resultAdFile = open(resultAdFileJson, 'w')
            resultAdFile.write(jsonResult)
            resultAdFile.close()
        else:
            resultAdFile = open(resultAdFileJson, 'a')
            resultAdFile.write(jsonResult)
            resultAdFile.close()
        self.insertListBoxMessage("生成成功")
        logging.info("生成成功")

    def creatMainUi(self):

        root = Tk()
        root.title("XPAD data pip json脚本生成工具")
        root.geometry("1000x618")
        root.resizable(False, False)

        self.path = StringVar()
        self.native_ad_new_usr_st = IntVar()
        self.native_ad_new_usr_st.set(253370736000000)

        self.native_ad_refr_it = IntVar()
        self.native_ad_refr_it.set(10000)
        self.native_ad_refr_t_u = IntVar()
        self.native_ad_refr_t_u.set(10)

        topFrame = Frame(root)
        topFrame.pack(side=TOP)

        Label(topFrame, text="目标路径:").pack(side=LEFT, padx=5, pady=10)
        Entry(topFrame, textvariable=self.path).pack(side=LEFT, padx=5, pady=10)
        ttk.Button(topFrame, text="路径选择", command=self.selectPath).pack(side=LEFT, padx=5, pady=10)
        ttk.Button(topFrame, text="开始生成", command=self.startCreateJson).pack(side=LEFT, padx=5, pady=10)

        middleFrame = Frame(root)
        middleFrame.pack(side=TOP)

        Label(middleFrame, text="新用户时间戳").pack(side=TOP, padx=5, pady=10)
        newUserAdFrame = Frame(middleFrame)
        newUserAdFrame.pack(side=TOP)
        Label(newUserAdFrame, text="ad_new_usr_st").pack(side=LEFT, padx=5, pady=10)
        Entry(newUserAdFrame, textvariable=self.native_ad_new_usr_st).pack(side=LEFT, padx=5, pady=10)

        Label(middleFrame, text="原生广告配置").pack(side=TOP, padx=5, pady=10)
        nativeAdFrame = Frame(middleFrame)
        nativeAdFrame.pack(side=TOP)

        Label(nativeAdFrame, text="ad_refr_it").pack(side=LEFT, padx=5, pady=10)
        Entry(nativeAdFrame, textvariable=self.native_ad_refr_it).pack(side=LEFT, padx=5, pady=10)
        Label(nativeAdFrame, text="ad_refr_t_u").pack(side=LEFT, padx=5, pady=10)
        Entry(nativeAdFrame, textvariable=self.native_ad_refr_t_u).pack(side=LEFT, padx=5, pady=10)

        bottomFrame = Frame(root)
        scrollbar = Scrollbar(bottomFrame)
        scrollbar.pack(side=RIGHT, fill=Y)
        self.listBox = Listbox(root, yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.listBox.yview)
        self.listBox.pack(side=LEFT, fill=BOTH, expand=YES, pady=10)
        bottomFrame.pack(side=TOP, fill=BOTH, expand=YES)

        self.insertListBoxMessage("欢迎来到XPAD XPAD data pip json脚本生成工具")
        self.insertListBoxMessage("请选择路径 :")

        root.mainloop()


if __name__ == '__main__':
    argLen = len(sys.argv)
    pip = XpadJsonBuildDataPip()
    if argLen < 2:
        pip.creatMainUi()
    else:
        excelPath = sys.argv[1]
        pip.startAndBuild(excelPath)
