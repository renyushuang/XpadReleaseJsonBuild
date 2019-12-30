# coding=utf-8
import openpyxl, os, json, sys
from openpyxl.worksheet.worksheet import Worksheet

AD_TYPE_OPEN = "开屏"
AD_TYPE_NATIVE_PLATE = "原生模版"
AD_TYPE_PLATE_RENDER = "模版渲染"
AD_TYPE_INTERSTITIAL = "插屏"
AD_TYPE_FULL_SCREEN_VIDEO = "全屏视频"
AD_TYPE_NATIVE = "原生"
AD_TYPE_NATIVE_FEED = "自渲染"
AD_TYPE_BANNER = "banner"

# extra-type
AD_TYPE_CSJ_NAME = "穿山甲"
AD_TYPE_CSJ = "csj"
# 穿山甲开屏
AD_TYPE_CSJ_OPEN: int = 11
# 穿山甲个性化模版banner
AD_TYPE_CSJ_PERSONAL_PLATE_BANNER: int = 12
# 穿山甲自渲染信息流
AD_TYPE_CSJ_FEED: int = 13
# 穿山甲个性化模版信息流
AD_TYPE_CSJ_PERSONAL_PLATE_FEED: int = 14
# 穿山甲全屏视频
AD_TYPE_CSJ_FULL_SCREEN_VIDEO: int = 15
# 穿山甲个性化模版插屏
AD_TYPE_CSJ_PERSONAL_PLATE_INTERSTITIAL: int = 16

# extra-type
AD_TYPE_YLH_NAME = "广点通"
AD_TYPE_YLH = "ylh"
# 优量汇开屏
AD_TYPE_YLH_OPEN: int = 21
# 优量汇banner2.0
AD_TYPE_YLH_BANNER: int = 22
# 优量汇插屏2.0
AD_TYPE_YLH_INTERSTITIAL: int = 23
# 优量汇原生模版
AD_TYPE_YLH_PERSONAL_PLATE_FEED: int = 24
# 优量汇原生自渲染
AD_TYPE_YLH_FEED: int = 25

# extra-type
AD_TYPE_KSH_NAME = "快手"
AD_TYPE_KSH = "ksh"
# 快手原生自渲染
AD_TYPE_KSH_FEED: int = 31
# 快手模版信息流
AD_TYPE_KSH_PERSONAL_PLATE_FEED: int = 32
# 快手全屏视频
AD_TYPE_KSH_FULL_SCREEN_VIDEO: int = 33

excelPath = ""

mAdResultMap = {}


def findTitleInColum(name, adSheet):
    maxColumn = int(adSheet.max_column)
    for columnIndex in range(0, maxColumn):
        titleValue = adSheet.cell(row=1, column=columnIndex + 1).value
        if titleValue == name:
            return columnIndex + 1

    print("title 不存在 --" + name)
    exit()


def getTitleColumValue(name, adSheet):
    index = findTitleInColum(name, adSheet)
    value = adSheet.cell(row=2, column=index).value
    if value is not None:
        return value

    print("当前 -- " + name + "值为 None")

    return ""


def getCloumeValueColumValue(row, name, adSheet: Worksheet):
    index = findTitleInColum(name, adSheet)
    value = adSheet.cell(row=row, column=index).value

    return value


def hasAdTypeString(adDetailsValuelist: list, typeName):
    length = len(adDetailsValuelist)

    for i in range(0, length):
        if adDetailsValuelist[i] == typeName:
            return True

    return False


def getAdExtraTypeValue(platformValue, adDetailsValue, adSourceIdValue):
    adDetailsValuelist: list = adDetailsValue.split("-")
    if len(adDetailsValuelist) < 2:
        print("备注 命名错误" + str(adSourceIdValue))

    if platformValue == AD_TYPE_CSJ:
        if hasAdTypeString(adDetailsValuelist, AD_TYPE_OPEN):
            return AD_TYPE_CSJ_OPEN
        elif hasAdTypeString(adDetailsValuelist, AD_TYPE_INTERSTITIAL):
            return AD_TYPE_CSJ_PERSONAL_PLATE_INTERSTITIAL
        elif hasAdTypeString(adDetailsValuelist, AD_TYPE_FULL_SCREEN_VIDEO):
            return AD_TYPE_CSJ_FULL_SCREEN_VIDEO
        elif hasAdTypeString(adDetailsValuelist, AD_TYPE_PLATE_RENDER):
            return AD_TYPE_CSJ_PERSONAL_PLATE_FEED
        elif hasAdTypeString(adDetailsValuelist, AD_TYPE_NATIVE_FEED):
            return AD_TYPE_CSJ_FEED
        elif hasAdTypeString(adDetailsValuelist, AD_TYPE_BANNER):
            return AD_TYPE_CSJ_PERSONAL_PLATE_BANNER
        else:
            print("不支持这种类型 " + adDetailsValue + "穿山甲的广告id为 = " + str(adSourceIdValue))

    elif platformValue == AD_TYPE_YLH:
        if hasAdTypeString(adDetailsValuelist, AD_TYPE_OPEN):
            return AD_TYPE_YLH_OPEN
        if hasAdTypeString(adDetailsValuelist, AD_TYPE_NATIVE_PLATE):
            return AD_TYPE_YLH_PERSONAL_PLATE_FEED
        if hasAdTypeString(adDetailsValuelist, AD_TYPE_INTERSTITIAL):
            return AD_TYPE_YLH_INTERSTITIAL
        if hasAdTypeString(adDetailsValuelist, AD_TYPE_NATIVE):
            return AD_TYPE_YLH_FEED
        if hasAdTypeString(adDetailsValuelist, AD_TYPE_BANNER):
            return AD_TYPE_YLH_BANNER
        else:
            print("不支持这种类型 " + adDetailsValue + "广点通id为 = " + str(adSourceIdValue))

        pass
    elif platformValue == AD_TYPE_KSH:
        if hasAdTypeString(adDetailsValuelist, AD_TYPE_NATIVE):
            return AD_TYPE_KSH_FEED
        if hasAdTypeString(adDetailsValuelist, AD_TYPE_PLATE_RENDER):
            return AD_TYPE_KSH_PERSONAL_PLATE_FEED
        if hasAdTypeString(adDetailsValuelist, AD_TYPE_FULL_SCREEN_VIDEO) or hasAdTypeString(adDetailsValuelist,
                                                                                             AD_TYPE_INTERSTITIAL):
            return AD_TYPE_KSH_FULL_SCREEN_VIDEO
        if hasAdTypeString(adDetailsValuelist, AD_TYPE_NATIVE):
            return AD_TYPE_KSH_PERSONAL_PLATE_FEED
        else:
            print("不支持这种类型 " + adDetailsValue + "快手id为 = " + str(adSourceIdValue))

        pass
    else:
        print("没有这个类型 广告id是 = " + str(adSourceIdValue))
        exit()

    print("不支持这种类型 " + adDetailsValue + "id为 = " + str(adSourceIdValue))
    return None


def get_merged_cells_value(adSheet, row_index, col_index):
    merged = adSheet.merged_cells
    for (min_col, min_row, max_col, max_row) in merged:
        if (row_index >= min_row[1] and row_index <= max_row[1]):
            if (col_index >= min_col[1] and col_index <= max_col[1]):
                cell_value = adSheet.cell(min_row[1], min_col[1])
                # print('该单元格[%d,%d]属于合并单元格，值为[%s]' % (row_index, col_index, cell_value.value))
                return cell_value.value
    return None


def addChannelIds(slot: list, adSheet):
    maxColum = adSheet.max_row
    channels = {}
    channelDatas = []
    adPriorityList = []

    for sidRowIndex in range(1, maxColum):

        currentIndex = sidRowIndex + 1
        sidValue = getCloumeValueColumValue(currentIndex, "sid", adSheet)

        if sidValue is not None:
            channels = {}
            channelDatas = []
            adPriorityList = []
            channels["channels"] = channelDatas
            slot.append(channels)
            channels["sid"] = sidValue
        else:
            # 跳过不属于当前单元格的输出
            merged_value = get_merged_cells_value(adSheet, currentIndex, 4)
            if merged_value != channels["sid"]:
                continue

        platformValue = getCloumeValueColumValue(currentIndex, "Platform", adSheet)
        if platformValue is None:
            print("Platform is None --- sid = " + str(channels["sid"]) + "----- 行 = " + str(currentIndex))
            continue

        adSourceIdValue = getCloumeValueColumValue(currentIndex, "广告ID", adSheet)
        if adSourceIdValue is None:
            print("广告ID is None --- sid = " + str(channels["sid"]) + "----- 行 = " + str(currentIndex))
            continue

        adWatingTimeValue = getCloumeValueColumValue(currentIndex, "广告超时时间", adSheet)
        if adWatingTimeValue is None:
            adWatingTimeValue = 4000
        elif adWatingTimeValue < 1000:
            print("广告超时时间 是毫秒的 =" + str(adSourceIdValue) + "--对应 = " + adWatingTimeValue + " 是错误的")

        adExpriedTImeValue = getCloumeValueColumValue(currentIndex, "广告过期时间", adSheet)
        if adExpriedTImeValue is None:
            adExpriedTImeValue = 40
        elif adExpriedTImeValue < 0:
            print("广告过期时间 是分钟的 =" + str(adSourceIdValue) + "--对应 = " + adWatingTimeValue + " 是错误的")

        adDetailsValue = getCloumeValueColumValue(currentIndex, "备注", adSheet)
        if adDetailsValue is None:
            print("备注内容为None  跳过= " + str(adSourceIdValue))
            continue

        adExtraType = getAdExtraTypeValue(platformValue, adDetailsValue, adSourceIdValue)

        channelItem = {}
        channelItemExtra = {}

        adPriorityValue = getCloumeValueColumValue(currentIndex, "广告优先级", adSheet)

        if adPriorityValue is not None:
            adPriorityListlength = len(adPriorityList)

            if adPriorityListlength == 0:
                channelDatas.append(channelItem)
                adPriorityList.append(adPriorityValue)
            else:
                for i in range(0, adPriorityListlength):
                    if adPriorityValue <= adPriorityList[i]:
                        adPriorityList.insert(i, adPriorityValue)
                        channelDatas.insert(i, channelItem)
                        break
                    else:
                        # i+1 小于最长的长度，可以再向前进一格子
                        if (i + 1) < adPriorityListlength:
                            continue
                        else:
                            # i+1 超出了长度 数据等于排序长度
                            if len(channelDatas) <= adPriorityListlength:
                                adPriorityList.append(adPriorityValue)
                                channelDatas.append(channelItem)
                            else:
                                adPriorityList.append(adPriorityValue)
                                channelDatas.insert(i + 1, channelItem)

        else:
            channelDatas.append(channelItem)

        # print("优先级顺序 -- " + str(adPriorityList))

        channelItem["extra"] = channelItemExtra
        channelItem["channel"] = platformValue
        channelItem["pid"] = adSourceIdValue
        channelItem["wt"] = adWatingTimeValue
        channelItem["ttl"] = adExpriedTImeValue

        channelItemExtra["type"] = adExtraType
        if adExtraType == AD_TYPE_CSJ_PERSONAL_PLATE_INTERSTITIAL:
            channelItemExtra["image_ratio"] = "2:3"
        elif adExtraType == AD_TYPE_CSJ_PERSONAL_PLATE_FEED:
            channelItemExtra["dip_w"] = "300"
            channelItemExtra["dip_h"] = "250"
        elif adExtraType == AD_TYPE_YLH_PERSONAL_PLATE_FEED:
            channelItemExtra["pixel_w"] = "300"
            channelItemExtra["pixel_h"] = "250"

    findTitleInColum("广告位名称", adSheet)

    pass


def main():
    # 开始读取
    wb = openpyxl.load_workbook(excelPath)
    sheetNames = wb.sheetnames
    adSheet: Worksheet = wb[str(sheetNames[0])]

    maxColumn = int(adSheet.max_column)

    print("最大列 = " + str(maxColumn))
    print("最大行 = " + str(adSheet.max_row))

    appLicenseId = getTitleColumValue("app license id", adSheet)
    csjAppId = getTitleColumValue("穿山甲应用ID", adSheet)
    ylhAppId = getTitleColumValue("广点通应用ID", adSheet)
    kshAppId = getTitleColumValue("快手应用ID", adSheet)
    appId = getTitleColumValue("appId", adSheet)

    mAdResultMap["ls"] = appLicenseId
    mAdResultMap["rt"] = 20
    adResultDataMap = {}
    mAdResultMap["data"] = adResultDataMap

    adResultDataMap["ls"] = appLicenseId
    adResultDataMap["csj"] = csjAppId
    adResultDataMap["ylh"] = ylhAppId
    adResultDataMap["ksh"] = kshAppId
    adResultDataMap["appid"] = appId
    slot = []
    adResultDataMap["slot"] = slot

    addChannelIds(slot, adSheet)

    mAdResultMap["status"] = 1


if __name__ == '__main__':
    print("XPAD 2.0 json脚本生成工具")
    argLen = len(sys.argv)
    if argLen < 2:
        print("请输入想要解析的文件")
        exit()
    excelPath = sys.argv[1]

    if not os.path.exists(excelPath):
        print("需要解析的Excel文件 不存在")
        exit()

    fileSub = os.path.splitext(os.path.basename(excelPath))[1]
    if fileSub != ".xlsx":
        print("请输入正确的Excel文件->" + fileSub)
        exit()
    print("请确保需要解析的广告数据表在第一个...")
    print("开始生成 ...")

    main()

    jsonResult = json.dumps(mAdResultMap, indent=4, ensure_ascii=False)
    fileName = os.path.splitext(os.path.basename(excelPath))[0]
    fileDir = os.path.dirname(excelPath)
    resultAdFileJson = os.path.join(fileDir, fileName + "_xpad_ad_release_2.0.json")
    print(resultAdFileJson)

    if os.path.exists(resultAdFileJson):
        resultAdFile = open(resultAdFileJson, 'w')
        resultAdFile.write(jsonResult)
        resultAdFile.close()
    else:
        resultAdFile = open(resultAdFileJson, 'a')
        resultAdFile.write(jsonResult)
        resultAdFile.close()
    print("生成成功")
